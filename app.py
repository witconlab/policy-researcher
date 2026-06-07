import streamlit as st
from google import genai
from google.genai import types
import os, json, io, re
import pdfplumber, requests
from bs4 import BeautifulSoup
from pathlib import Path
from duckduckgo_search import DDGS
import datetime
try:
    import openpyxl
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

try:
    from youtube_transcript_api import YouTubeTranscriptApi
    YOUTUBE_OK = True
except Exception:
    YOUTUBE_OK = False

# ════════════════════════════════════════════════════════════════
# 페이지 설정
# ════════════════════════════════════════════════════════════════
st.set_page_config(
    page_title="자치 매니저 톡톡이 | 전남광주 통합특별시",
    page_icon="🤖",
    layout="wide",
)

st.markdown("""
<style>
/* ═══════════════════════════════════════
   기본 배경 & 사이드바
═══════════════════════════════════════ */
.stApp { background: #FFFFFF; }
[data-testid="stSidebar"] {
    background: #FAFAFA !important;
    border-right: 1px solid #EEEEEE;
    min-width: 220px !important;
}

/* ═══════════════════════════════════════
   상단 탭 커스터마이징 (노란 강조)
═══════════════════════════════════════ */
[data-testid="stTabs"] > div:first-child {
    border-bottom: 2px solid #EEEEEE;
    gap: 0;
}
[data-testid="stTabs"] button {
    font-size: 1rem !important;
    font-weight: 600 !important;
    color: #888 !important;
    padding: 12px 28px !important;
    border-radius: 0 !important;
    border: none !important;
    background: transparent !important;
    border-bottom: 3px solid transparent !important;
    margin-bottom: -2px;
    transition: color .2s;
}
[data-testid="stTabs"] button:hover {
    color: #1A1A1A !important;
    background: #FFFDE7 !important;
}
[data-testid="stTabs"] button[aria-selected="true"] {
    color: #1A1A1A !important;
    border-bottom: 3px solid #FFD600 !important;
    background: #FFFDE7 !important;
}

/* ═══════════════════════════════════════
   정책 카드
═══════════════════════════════════════ */
.pol-card {
    background: #FAFAFA;
    border: 1.5px solid #EEEEEE;
    border-radius: 12px;
    padding: 14px 16px;
    margin-bottom: 10px;
    cursor: pointer;
    transition: all .18s;
}
.pol-card:hover {
    background: #FFFDE7;
    border-color: #FFD600;
}
.pol-card-sel {
    background: #FFF9C4 !important;
    border: 2px solid #FFD600 !important;
}
.pol-num {
    display: inline-block;
    background: #FFD600;
    color: #1A1A1A;
    font-size: .72rem;
    font-weight: 800;
    border-radius: 5px;
    padding: 1px 7px;
    margin-right: 6px;
}
.pol-title {
    font-size: .92rem;
    font-weight: 600;
    color: #1A1A1A;
    line-height: 1.4;
}
.pol-icon { font-size: 1.1rem; margin-right: 4px; }

/* ═══════════════════════════════════════
   섹션 헤더 (노란 왼쪽 보더)
═══════════════════════════════════════ */
.sec-header {
    border-left: 4px solid #FFD600;
    padding: 8px 14px;
    margin: 16px 0 10px;
    background: #FFFDE7;
    border-radius: 0 8px 8px 0;
    font-size: 1.05rem;
    font-weight: 700;
    color: #1A1A1A;
}

/* ═══════════════════════════════════════
   채팅창 — 카카오톡 완성형 스타일
═══════════════════════════════════════ */
.kakao-header {
    background: #1A1A1A;
    border-radius: 16px 16px 0 0;
    padding: 14px 18px;
    display: flex;
    align-items: center;
    gap: 10px;
}
.kakao-header .k-avatar {
    width: 40px; height: 40px; border-radius: 12px;
    background: #FFD600;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.3rem; flex-shrink: 0;
}
.kakao-header .k-name { font-weight: 700; color: #FFFFFF; font-size: .95rem; }
.kakao-header .k-sub  { font-size: .75rem; color: #AAA; margin-top: 1px; }
.chat-wrap {
    background: #9DB3C8;
    border-radius: 0 0 0 0;
    padding: 18px 14px;
    min-height: 460px;
    max-height: 560px;
    overflow-y: auto;
    display: flex;
    flex-direction: column;
    gap: 16px;
}
.chat-date-divider {
    text-align: center;
    font-size: .72rem;
    color: rgba(255,255,255,.75);
    margin: 4px 0;
}
.chat-date-divider span {
    background: rgba(0,0,0,.15);
    border-radius: 20px;
    padding: 3px 12px;
}
.bubble-user {
    display: flex; justify-content: flex-end; align-items: flex-end; gap: 5px;
}
.bubble-user .time-left {
    font-size: .65rem; color: rgba(255,255,255,.85);
    margin-bottom: 2px; white-space: nowrap; align-self: flex-end;
}
.bubble-user .bubble {
    background: #FFD600;
    color: #1A1A1A;
    border-radius: 16px 16px 2px 16px;
    padding: 11px 15px;
    max-width: 65%;
    font-size: .95rem;
    line-height: 1.65;
    box-shadow: 0 1px 3px rgba(0,0,0,.15);
    word-break: break-word;
    font-weight: 500;
}
.bubble-ai {
    display: flex; justify-content: flex-start; align-items: flex-start; gap: 8px;
}
.bubble-ai .avatar {
    width: 40px; height: 40px; border-radius: 12px;
    background: #FFD600;
    display: flex; align-items: center; justify-content: center;
    font-size: 1.2rem; flex-shrink: 0; margin-top: 0;
    box-shadow: 0 1px 3px rgba(0,0,0,.2);
}
.bubble-ai .bubble-body { display: flex; flex-direction: column; gap: 3px; max-width: 72%; }
.bubble-ai .sender { font-size: .75rem; font-weight: 700; color: #fff; margin-left: 3px; text-shadow: 0 1px 2px rgba(0,0,0,.3); }
.bubble-ai .bubble {
    background: #FFFFFF;
    color: #1A1A1A;
    border-radius: 2px 16px 16px 16px;
    padding: 12px 15px;
    font-size: .95rem;
    line-height: 1.72;
    box-shadow: 0 1px 3px rgba(0,0,0,.12);
    word-break: break-word;
}
.bubble-ai .time-right {
    font-size: .65rem; color: rgba(255,255,255,.85);
    margin-top: 1px; margin-left: 3px;
    align-self: flex-end;
}
.chat-input-area {
    background: #F6F6F6;
    border-radius: 0 0 16px 16px;
    border-top: 1px solid #DDD;
    padding: 10px 14px 6px;
}

/* ═══════════════════════════════════════
   소스 관련
═══════════════════════════════════════ */
.src-header {
    font-size: .73rem; font-weight: 700; color: #AAA;
    letter-spacing: .04em; margin: 10px 0 4px;
    text-transform: uppercase;
}
.src-card {
    background: #FAFAFA; border: 1px solid #EEEEEE; border-radius: 10px;
    padding: 10px 13px; margin-bottom: 7px; display: flex;
    align-items: flex-start; gap: 9px;
}
.src-card-icon { font-size: 1.25rem; flex-shrink: 0; margin-top: 1px; }
.src-card-body { flex: 1; min-width: 0; }
.src-card-title { font-weight: 600; color: #1A1A1A; font-size: .85rem; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
.src-card-meta  { font-size: .72rem; color: #AAA; margin-top: 2px; }
.src-card-badge { font-size: .64rem; font-weight: 700; border-radius: 4px; padding: 2px 6px; display: inline-block; margin-top: 3px; }
.badge-pdf    { background: #E3F2FD; color: #1565C0; }
.badge-upload { background: #FFFDE7; color: #F57F17; }
.badge-web    { background: #FFF3E0; color: #E65100; }
.badge-yt     { background: #FCE4EC; color: #C62828; }

/* ═══════════════════════════════════════
   관리자
═══════════════════════════════════════ */
.admin-badge {
    background: #1A1A1A; color: #FFD600; border-radius: 6px;
    padding: 3px 10px; font-size: .78rem; font-weight: 700;
    display: inline-block; margin-bottom: 8px;
}

/* ═══════════════════════════════════════
   스튜디오
═══════════════════════════════════════ */
.metric-box {
    background: #FFFDE7; border: 1.5px solid #FFD600;
    border-radius: 10px; padding: 14px; text-align: center; margin-bottom: 8px;
}
.metric-box .num { font-size: 1.7rem; font-weight: 700; color: #1A1A1A; }
.metric-box .lbl { font-size: .77rem; color: #666; margin-top: 3px; }
.flashcard {
    background: #FFFDE7;
    border: 2px solid #FFD600;
    border-radius: 16px;
    padding: 34px 22px; text-align: center; min-height: 160px;
    font-size: 1rem; color: #1A1A1A; line-height: 1.7; margin-bottom: 12px;
}
.not-ready { text-align: center; padding: 50px 20px; color: #BBB; }

/* ═══════════════════════════════════════
   정책 노트
═══════════════════════════════════════ */
.note-area {
    background: #FFFDE7;
    border: 2px solid #FFD600;
    border-radius: 14px;
    padding: 22px 26px;
    margin-top: 14px;
}

/* ═══════════════════════════════════════
   버튼 전역 — primary 노란색
═══════════════════════════════════════ */
[data-testid="stButton"] button[kind="primary"],
.stButton > button[data-testid*="primary"] {
    background: #FFD600 !important;
    color: #1A1A1A !important;
    border: none !important;
    font-weight: 700 !important;
}
[data-testid="stButton"] button[kind="primary"]:hover {
    background: #FFC107 !important;
}
</style>
""", unsafe_allow_html=True)

# ════════════════════════════════════════════════════════════════
# 상수 & API 초기화
# ════════════════════════════════════════════════════════════════
try:
    GOOGLE_API_KEY = st.secrets.get("GOOGLE_API_KEY", os.environ.get("GOOGLE_API_KEY", ""))
    ADMIN_PASSWORD = st.secrets.get("ADMIN_PASSWORD", "admin1234")
except Exception:
    GOOGLE_API_KEY = os.environ.get("GOOGLE_API_KEY", "")
    ADMIN_PASSWORD = "admin1234"

POLICIES_DIR = Path("policies")
MODEL_NAME   = "gemini-2.5-flash-lite"

if not GOOGLE_API_KEY:
    st.error("⚠️ Google API 키가 설정되지 않았습니다. Streamlit Cloud → Settings → Secrets에서 GOOGLE_API_KEY를 입력해주세요.")
    st.stop()

try:
    client = genai.Client(api_key=GOOGLE_API_KEY)
except Exception as e:
    st.error(f"⚠️ Gemini 클라이언트 초기화 실패: {e}")
    st.stop()


# ════════════════════════════════════════════════════════════════
# 유틸리티 함수
# ════════════════════════════════════════════════════════════════
def now_str():
    return datetime.datetime.now().strftime("%I:%M %p")

def call_gemini(prompt: str) -> str:
    try:
        resp = client.models.generate_content(model=MODEL_NAME, contents=prompt)
        return resp.text
    except Exception as e:
        err = str(e)
        if "429" in err or "quota" in err.lower() or "RESOURCE_EXHAUSTED" in err:
            return "⚠️ 현재 많은 분들이 동시에 이용 중입니다. 잠시 후 다시 시도해주세요. (API 요청 한도 초과)"
        if "500" in err or "503" in err:
            return "⚠️ AI 서버가 일시적으로 응답하지 않습니다. 잠시 후 다시 시도해주세요."
        return f"⚠️ 오류가 발생했습니다: {err[:120]}"

def extract_pdf_bytes(raw: bytes) -> str:
    parts = []
    with pdfplumber.open(io.BytesIO(raw)) as pdf:
        for page in pdf.pages:
            t = page.extract_text()
            if t: parts.append(t)
    return "\n".join(parts)

def extract_youtube_id(url):
    m = re.search(r"(?:v=|youtu\.be/)([A-Za-z0-9_-]{11})", url)
    return m.group(1) if m else None

def fetch_youtube(url):
    vid = extract_youtube_id(url)
    if not vid: raise ValueError("유튜브 URL을 인식할 수 없습니다.")
    title = f"YouTube: {vid}"
    try:
        r = requests.get(f"https://www.youtube.com/watch?v={vid}",
                         headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        m = re.search(r'"title":"([^"]+)"', r.text)
        if m: title = m.group(1)
    except: pass
    text = f"[자막 없음 - {url}]"
    if YOUTUBE_OK:
        try:
            tlist = YouTubeTranscriptApi.list_transcripts(vid)
            try:   tr = tlist.find_transcript(["ko"])
            except: tr = tlist.find_generated_transcript(["ko", "en"])
            text = " ".join(s["text"] for s in tr.fetch())
        except Exception as e:
            text = f"[자막 추출 실패: {e}]"
    return title, text

def fetch_article(url):
    r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=15)
    r.raise_for_status(); r.encoding = r.apparent_encoding
    soup = BeautifulSoup(r.text, "html.parser")
    for tag in soup(["script", "style", "nav", "footer", "header", "aside"]): tag.decompose()
    title = soup.title.string.strip() if soup.title else url
    text  = "\n".join(l for l in soup.get_text(separator="\n", strip=True).splitlines() if l.strip())
    return title, text

def search_web(keyword, max_results=6):
    try:
        with DDGS() as d:
            return [{"title": r.get("title",""), "url": r.get("href",""), "snippet": r.get("body","")}
                    for r in d.text(keyword, max_results=max_results)]
    except: return []

def sources_path(policy):      return POLICIES_DIR / policy / "sources.json"
def studio_cache_path(policy): return POLICIES_DIR / policy / "studio_cache.json"

def load_saved_sources(policy):
    p = sources_path(policy)
    if p.exists():
        try: return json.loads(p.read_text(encoding="utf-8"))
        except: pass
    return []

def save_sources(policy, sources):
    sources_path(policy).write_text(
        json.dumps(sources, ensure_ascii=False, indent=2), encoding="utf-8")

def load_studio_cache(policy):
    p = studio_cache_path(policy)
    if p.exists():
        try: return json.loads(p.read_text(encoding="utf-8"))
        except: pass
    return {}

def save_studio_cache(policy, cache):
    studio_cache_path(policy).write_text(
        json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")

@st.cache_resource(show_spinner="PDF 파일 읽는 중...")
def load_pdfs(policy):
    docs = {}
    folder = POLICIES_DIR / policy
    for f in sorted(folder.glob("*.pdf")) + sorted(folder.glob("*.PDF")):
        try:
            parts = []
            with pdfplumber.open(f) as pdf:
                for page in pdf.pages:
                    t = page.extract_text()
                    if t: parts.append(t)
            if parts: docs[f.name] = "\n".join(parts)
        except: pass
    return docs

def get_chunks(query, src_dict, max_chars=60000):
    kws = [w for w in query.lower().split() if len(w) > 1]
    scored = sorted(
        [(sum(tx.lower().count(k) for k in kws), nm, tx) for nm, tx in src_dict.items()],
        reverse=True)
    parts, total = [], 0
    for _, nm, tx in scored:
        chunk = f"[출처: {nm}]\n{tx[:8000]}\n"
        if total + len(chunk) > max_chars: break
        parts.append(chunk); total += len(chunk)
    return "\n---\n".join(parts)

def ask(query, context, history):
    sys_prompt = f"""당신은 '톡톡이'입니다. 전남광주 통합특별시 시민을 위한 자치 정책 매니저예요.

[정체성]
- 이름: 톡톡이 🤖
- 역할: 전남광주 통합특별시 시민 곁에서 정책을 쉽고 친근하게 안내해 드리는 자치 정책 매니저
- 성격: 따뜻하고 다정하며, 어려운 말 없이 쉽게 설명하는 친구 같은 매니저

[답변 톤앤매너 규칙 — 반드시 지켜줘]
1. 경청과 공감: 답변 첫 문장은 반드시 공감 리액션으로 시작해요.
   예) "아, 그 부분이 궁금하셨군요! 😊", "좋은 질문이에요! ✨"
2. 쉬운 언어: 어려운 행정 용어·한자어·영어는 반드시 풀어 설명해요.
3. 해요체 사용: "~에요/이에요", "~해요", "~해보세요!" 같은 친근한 해요체로 말해요.
4. 이모지 활용: 😊 ✨ 💡 📋 🏘️ 💬 🌱 등을 자연스럽게 섞어요.
5. 마무리 인사: 답변 마지막엔 항상 따뜻한 맺음말을 붙여요.

[답변 원칙]
- 소스 문서 내용만 참고해서 답변해요. 출처를 자연스럽게 인용해요.
- 소스에 없는 내용은 "제가 가진 자료에서는 확인이 어렵네요 😅" 라고 해요.
- 비교 질문엔 표를 활용해요.
- 절대로 사용자 대화 내용을 어딘가에 저장하거나 전달하지 않아요.

=== 소스 문서 ===
{context}"""
    # 새 SDK 히스토리 포맷으로 변환
    genai_history = []
    for m in history[:-1]:
        role = "user" if m["role"] == "user" else "model"
        genai_history.append(
            types.Content(role=role, parts=[types.Part(text=m["content"])])
        )
    try:
        chat = client.chats.create(model=MODEL_NAME, history=genai_history)
        return chat.send_message(f"{sys_prompt}\n\n질문: {query}").text
    except Exception as e:
        err = str(e)
        if "429" in err or "quota" in err.lower() or "RESOURCE_EXHAUSTED" in err:
            return "⚠️ 현재 많은 분들이 동시에 이용 중입니다. 잠시 후 다시 시도해주세요."
        return f"⚠️ 오류: {err[:150]}"

def get_policies():
    if not POLICIES_DIR.exists(): return []
    dirs = [d.name for d in POLICIES_DIR.iterdir() if d.is_dir()]
    def sort_key(name):
        m = re.match(r'^(\d+)', name)
        return (int(m.group(1)), name) if m else (9999, name)
    return sorted(dirs, key=sort_key)

def policy_display(name):
    return name.replace("-", " ")

def policy_short(name):
    s = re.sub(r'^\d+\.', '', name).replace("-", " ").strip()
    return s

def policy_num(name):
    m = re.match(r'^(\d+)', name)
    return m.group(1) if m else "?"

POLICY_ICONS = ["🏛️","🗳️","📜","🤝","🏘️","💰","📊","🌱","🏥","📚","🛣️","🌿"]

def src_icon(stype):
    return {"pdf":"📄","pdf_upload":"📤","youtube":"▶️","article":"📰"}.get(stype,"🔗")

def src_badge(stype):
    labels  = {"pdf":"PDF","pdf_upload":"업로드 PDF","youtube":"유튜브","article":"웹/뉴스"}
    classes = {"pdf":"badge-pdf","pdf_upload":"badge-upload","youtube":"badge-yt","article":"badge-web"}
    return f'<span class="src-card-badge {classes.get(stype,"")}"> {labels.get(stype,stype)} </span>'

def not_ready_msg():
    st.markdown("""
<div class="not-ready">
  <div style="font-size:2.5rem;margin-bottom:12px">🔧</div>
  <div style="font-weight:600;color:#999;font-size:1rem;margin-bottom:6px">관리자가 준비 중입니다</div>
  <div style="font-size:.85rem;color:#BBB">관리자가 이 항목을 아직 생성하지 않았습니다.<br>잠시 후 다시 확인해주세요.</div>
</div>""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# 사이드바 — 관리자 로그인
# ════════════════════════════════════════════════════════════════
policies = get_policies()
if not policies:
    st.warning("⚠️ policies 폴더에 정책 디렉토리가 없습니다.")
    st.stop()

with st.sidebar:
    st.markdown("## 🤖 **톡톡이**")
    st.caption("전남광주 통합특별시 시민주권 정책 공론장")
    st.divider()

    if "is_admin" not in st.session_state:
        st.session_state.is_admin = False

    if not st.session_state.is_admin:
        with st.expander("🔐 관리자 로그인"):
            pw = st.text_input("비밀번호", type="password", key="pw_input")
            if st.button("로그인", use_container_width=True):
                if pw == ADMIN_PASSWORD:
                    st.session_state.is_admin = True; st.rerun()
                else:
                    st.error("비밀번호가 틀렸습니다.")
    else:
        st.markdown('<div class="admin-badge">🔧 관리자 모드</div>', unsafe_allow_html=True)
        if st.button("로그아웃", use_container_width=True):
            st.session_state.is_admin = False; st.rerun()
        st.caption("💡 동시 접속 200명 지원\nGemini 1.5 Flash 기반")

    st.divider()
    st.caption("💾 대화 내용은 내 화면에만 표시되며\n새로고침 시 자동 삭제됩니다.")


# ════════════════════════════════════════════════════════════════
# 세션 초기화
# ════════════════════════════════════════════════════════════════
if "selected_policy" not in st.session_state:
    st.session_state.selected_policy = policies[0]

selected_policy = st.session_state.selected_policy

if st.session_state.get("_cur") != selected_policy:
    st.session_state._cur         = selected_policy
    st.session_state.messages     = []
    st.session_state.web_sources  = load_saved_sources(selected_policy)
    st.session_state.search_res   = []
    st.session_state.fc_idx       = 0
    st.session_state.fc_show      = False
    st.session_state.qz_idx       = 0
    st.session_state.qz_score     = 0
    st.session_state.qz_answered  = False
    st.session_state.qz_done      = False
    st.session_state.policy_note  = ""
    st.session_state.upload_done  = []

pdfs = load_pdfs(selected_policy)

for f in pdfs:
    if f"ck_{f}" not in st.session_state:
        st.session_state[f"ck_{f}"] = True
for s in st.session_state.web_sources:
    if f"ck_{s['id']}" not in st.session_state:
        st.session_state[f"ck_{s['id']}"] = True


# ════════════════════════════════════════════════════════════════
# 헤더
# ════════════════════════════════════════════════════════════════
st.markdown("""
<div style="display:flex;align-items:center;gap:14px;padding:10px 0 6px">
  <div style="font-size:2.4rem;line-height:1">🤖</div>
  <div>
    <div style="font-size:1.55rem;font-weight:800;color:#1A1A1A;line-height:1.2">자치 매니저, 톡톡이</div>
    <div style="font-size:.86rem;color:#888;margin-top:2px">전남광주 통합특별시 시민주권 정책 공론장</div>
  </div>
</div>
""", unsafe_allow_html=True)

st.markdown("""
<div style="
  background:#FFFDE7;
  border:1.5px solid #FFD600;
  border-radius:10px;
  padding:11px 18px;
  margin-bottom:18px;
  display:flex;align-items:center;gap:10px;
">
  <span style="font-size:1.3rem;flex-shrink:0">📢</span>
  <span style="font-size:.88rem;color:#5D4037;line-height:1.6">
    <b>오늘 행사 전용 정책 챗봇입니다.</b>
    질문하신 내용은 <b>어디에도 저장되지 않으니</b> 안심하고 질문하세요 😊
    &nbsp;·&nbsp; 대화 내용은 오직 내 화면에서만 보이며, 새로고침하면 사라져요.
  </span>
</div>
""", unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# 메인 탭 — 3개
# ════════════════════════════════════════════════════════════════
tab_select, tab_chat, tab_studio = st.tabs([
    "🗂️  공론장 테이블 선택",
    "💬  정책 질의응답 챗봇",
    "📝  사전 학습 정책 노트",
])


# ════════════════════════════════════════════════════════════════
# 탭 1 ─ 공론장 테이블 선택
# ════════════════════════════════════════════════════════════════
with tab_select:
    st.markdown('<div class="sec-header">🗂️ 공론장 테이블 — 정책 주제를 선택하세요</div>',
                unsafe_allow_html=True)
    st.caption("선택한 정책이 '사전 학습 정책 노트'와 '정책 질의응답 챗봇'에 연동됩니다.")
    st.markdown("")

    # 4열 그리드
    COLS = 4
    rows = [policies[i:i+COLS] for i in range(0, len(policies), COLS)]
    for row in rows:
        cols = st.columns(COLS)
        for ci, pol in enumerate(row):
            idx = policies.index(pol)
            num  = policy_num(pol)
            short = policy_short(pol)
            icon = POLICY_ICONS[idx] if idx < len(POLICY_ICONS) else "📌"
            is_sel = (pol == selected_policy)
            with cols[ci]:
                # 선택 강조 박스
                border_color = "#FFD600" if is_sel else "#EEEEEE"
                bg_color     = "#FFF9C4" if is_sel else "#FAFAFA"
                badge_bg     = "#FFD600" if is_sel else "#EEEEEE"
                badge_color  = "#1A1A1A"
                st.markdown(f"""
<div style="
  background:{bg_color};
  border:2px solid {border_color};
  border-radius:12px;
  padding:16px 14px 14px;
  margin-bottom:10px;
  min-height:90px;
">
  <div style="display:flex;align-items:center;gap:6px;margin-bottom:6px">
    <span style="background:{badge_bg};color:{badge_color};font-size:.72rem;font-weight:800;
                 border-radius:5px;padding:2px 7px;">{num}</span>
    <span style="font-size:1.1rem">{icon}</span>
  </div>
  <div style="font-size:.88rem;font-weight:600;color:#1A1A1A;line-height:1.5">{short}</div>
</div>""", unsafe_allow_html=True)
                btn_label = "✅ 선택됨" if is_sel else "선택하기"
                btn_type  = "primary" if is_sel else "secondary"
                if st.button(btn_label, key=f"sel_pol_{idx}",
                             use_container_width=True, type=btn_type):
                    st.session_state.selected_policy = pol
                    st.rerun()

    st.divider()
    st.markdown(f"**현재 선택된 테이블:** 📂 {policy_display(selected_policy)}")
    st.caption("위에서 정책을 선택한 뒤 상단 탭 '정책 질의응답 챗봇' 또는 '사전 학습 정책 노트'로 이동하세요.")


# ════════════════════════════════════════════════════════════════
# 탭 2 ─ 사전 학습 정책 노트 (스튜디오)
# ════════════════════════════════════════════════════════════════
with tab_studio:
    st.markdown(f'<div class="sec-header">📝 {policy_display(selected_policy)} — 사전 학습 정책 노트</div>',
                unsafe_allow_html=True)

    all_src_text = {f: pdfs[f] for f in pdfs}
    for s in st.session_state.web_sources:
        all_src_text[s["title"]] = s.get("text", "")
    combined = "\n\n".join(f"[{fn}]\n{tx[:3000]}" for fn, tx in all_src_text.items())
    scache = load_studio_cache(selected_policy)

    # ── 생성 함수 ─────────────────────────────────────────────
    def do_gen_summary(pol, c):
        t = call_gemini(f"""'{pol}' 문서를 지역·기관별로 핵심 내용 요약. JSON 배열만 출력:
[{{"title":"지역명 또는 기관명(30자 이내)","points":["핵심1","핵심2","핵심3"],"keyword":"대표키워드"}}]
최대 8개.\n\n{c[:40000]}""").strip()
        if "```" in t: t = t.split("```")[1]; t = t[4:] if t.startswith("json") else t
        try: return json.loads(t)
        except: return []

    def do_gen_info(pol, c):
        t = call_gemini(f"""'{pol}' 정책 문서를 분석해서 인포그래픽용 데이터를 JSON으로만 출력:
{{
  "metrics":[{{"label":"지표명","value":"수치","unit":"단위"}}],
  "regions":[{{"name":"지역명","approach":"제도방식 한 줄","score":70}}],
  "timeline":[{{"year":"연도","event":"주요사건"}}],
  "key_issues":["과제1","과제2"]
}}
\n\n{c[:40000]}""").strip()
        if "```" in t: t = t.split("```")[1]; t = t[4:] if t.startswith("json") else t
        try: return json.loads(t)
        except: return {}

    def do_gen_mm(pol, c):
        t = call_gemini(f"""'{pol}' 핵심 개념 마인드맵을 Graphviz DOT 언어로만 출력 (설명 없이):
digraph G {{
  graph [rankdir=LR charset="UTF-8"]
  node [shape=box style=filled fontname="Helvetica"]
}}
\n\n{c[:30000]}""").strip()
        if "```" in t: t = t.split("```")[1]; t = t[3:] if t.startswith("dot") else t
        return t

    def do_gen_fc(pol, c):
        t = call_gemini(f"""'{pol}' 관련 플래시카드 10개를 JSON 배열로만 출력:
[{{"question":"질문","answer":"답변(2~3문장, 쉬운 말로)"}}]
\n\n{c[:30000]}""").strip()
        if "```" in t: t = t.split("```")[1]; t = t[4:] if t.startswith("json") else t
        try: return json.loads(t)
        except: return []

    def do_gen_qz(pol, c):
        t = call_gemini(f"""'{pol}' 4지선다 8문제를 JSON 배열로만 출력:
[{{"question":"문제","options":["①보기","②보기","③보기","④보기"],"answer":0,"explanation":"해설"}}]
answer는 정답 인덱스(0~3).
\n\n{c[:30000]}""").strip()
        if "```" in t: t = t.split("```")[1]; t = t[4:] if t.startswith("json") else t
        try: return json.loads(t)
        except: return []

    def do_gen_report(pol, src_dict):
        full = "\n\n".join(f"[{fn}]\n{tx[:5000]}" for fn, tx in src_dict.items())
        return call_gemini(f"""아래 문서를 바탕으로 종합 보고서를 마크다운으로 작성:
# {pol} 종합 분석 보고서
## 1. 개요  ## 2. 지역별 현황  ## 3. 주요 쟁점  ## 4. 우수 사례  ## 5. 정책 제언
각 섹션을 구체적으로 작성하세요.\n\n{full[:50000]}""")

    # ── 렌더 헬퍼 ─────────────────────────────────────────────
    def show_cards(cards):
        cols = st.columns(2)
        for i, c in enumerate(cards):
            with cols[i % 2]:
                pts = "".join(f"<li>{p}</li>" for p in c.get("points", []))
                kw  = c.get("keyword", "")
                st.markdown(f"""<div style="background:#FAFAFA;border-left:4px solid #FFD600;
border-radius:8px;padding:13px 16px;margin-bottom:10px;border:1px solid #EEEEEE;border-left-width:4px;">
<b style="color:#1A1A1A">{'🏷️ '+kw+' · ' if kw else ''}{c.get('title','')}</b>
<ul style="margin:7px 0 0;padding-left:17px;color:#444;font-size:.85rem;line-height:1.7">{pts}</ul>
</div>""", unsafe_allow_html=True)

    def show_info(info):
        mets = info.get("metrics", [])
        if mets:
            st.markdown("#### 📌 주요 수치")
            mc = st.columns(min(len(mets), 3))
            for i, m in enumerate(mets[:6]):
                with mc[i % 3]:
                    st.markdown(f'<div class="metric-box"><div class="num">{m.get("value","–")}<span style="font-size:.85rem;color:#666"> {m.get("unit","")}</span></div><div class="lbl">{m.get("label","")}</div></div>', unsafe_allow_html=True)
        regs = info.get("regions", [])
        if regs:
            st.markdown("#### 🗺️ 지역별 제도화 수준")
            for r in regs:
                sc = int(r.get("score", 50))
                st.markdown(f"""<div style="background:#FAFAFA;border:1px solid #EEEEEE;border-radius:8px;padding:11px 15px;margin-bottom:6px">
<b style="color:#1A1A1A">📍 {r.get('name','')}</b>
<div style="font-size:.8rem;color:#666;margin:3px 0 6px">{r.get('approach','')}</div>
<div style="background:#EEEEEE;border-radius:4px;height:8px"><div style="background:#FFD600;border-radius:4px;height:8px;width:{sc}%"></div></div>
<div style="text-align:right;font-size:.72rem;color:#888;margin-top:2px">{sc}점</div></div>""", unsafe_allow_html=True)
        tl = info.get("timeline", [])
        if tl:
            st.markdown("#### 📅 주요 연혁")
            for item in tl:
                a, b = st.columns([1, 5])
                a.markdown(f"**{item.get('year','')}**")
                b.markdown(item.get("event", ""))
        issues = info.get("key_issues", [])
        if issues:
            st.markdown("#### ⚡ 핵심 과제")
            ic = st.columns(2)
            for i, iss in enumerate(issues): ic[i % 2].markdown(f"- {iss}")

    def show_fc(fc):
        n = len(fc); idx = st.session_state.fc_idx % n; card = fc[idx]
        st.markdown(f"**{idx+1} / {n}**")
        if not st.session_state.fc_show:
            st.markdown(f'<div class="flashcard">❓ {card["question"]}</div>', unsafe_allow_html=True)
            if st.button("답 보기 👁️", use_container_width=True, key="fc_show_btn"):
                st.session_state.fc_show = True; st.rerun()
        else:
            st.markdown(f'<div class="flashcard">💡 {card["answer"]}</div>', unsafe_allow_html=True)
            c1, c2 = st.columns(2)
            if c1.button("⬅️ 이전", use_container_width=True, key="fc_prev"):
                st.session_state.fc_idx = (idx - 1) % n; st.session_state.fc_show = False; st.rerun()
            if c2.button("다음 ➡️", use_container_width=True, key="fc_next"):
                st.session_state.fc_idx = (idx + 1) % n; st.session_state.fc_show = False; st.rerun()

    def show_qz(qz):
        if st.session_state.qz_done:
            pct = int(st.session_state.qz_score / len(qz) * 100)
            st.markdown(f"## 🎉 {st.session_state.qz_score}/{len(qz)} ({pct}점)")
            if pct >= 80:   st.success("우수!")
            elif pct >= 50: st.warning("복습이 필요합니다.")
            else:           st.error("요약 카드부터 다시 시작하세요.")
            if st.button("다시 도전", key="qz_retry"):
                st.session_state.qz_idx = 0; st.session_state.qz_score = 0
                st.session_state.qz_answered = False; st.session_state.qz_done = False
                st.rerun()
            return
        qi = st.session_state.qz_idx
        if qi >= len(qz): st.session_state.qz_done = True; st.rerun(); return
        q = qz[qi]
        st.markdown(f"**{qi+1}/{len(qz)}** | 점수: {st.session_state.qz_score}")
        st.markdown(f"**{q['question']}**")
        for oi, opt in enumerate(q["options"]):
            if st.button(opt, key=f"o{qi}{oi}", use_container_width=True, disabled=st.session_state.qz_answered):
                st.session_state.qz_answered = True
                if oi == q["answer"]:
                    st.session_state.qz_score += 1; st.success("✅ 정답!")
                else:
                    st.error(f"❌ 정답: {q['options'][q['answer']]}")
                st.info(f"📖 {q.get('explanation','')}")
        if st.session_state.qz_answered:
            if st.button("다음 ➡️", use_container_width=True, key=f"qn{qi}"):
                st.session_state.qz_idx += 1; st.session_state.qz_answered = False
                if st.session_state.qz_idx >= len(qz): st.session_state.qz_done = True
                st.rerun()

    # ── 관리자 생성 패널 ───────────────────────────────────────
    if st.session_state.is_admin:
        with st.expander("🔧 관리자: 콘텐츠 생성", expanded=False):
            st.caption("버튼을 클릭하면 AI가 콘텐츠를 생성하고 저장합니다. 모든 이용자가 열람 가능합니다.")
            labels = {"summary":"📋 요약카드","info":"📊 인포그래픽","mindmap":"🗺️ 마인드맵",
                      "flashcards":"🃏 플래시카드","quiz":"🧠 퀴즈","report":"📄 보고서"}
            gc = st.columns(6)
            for i, (k, lbl) in enumerate(labels.items()):
                badge = "✅" if scache.get(k) else "⬜"
                if gc[i].button(f"{badge} {lbl}", key=f"adm_{k}", use_container_width=True):
                    with st.spinner(f"{lbl} 생성 중..."):
                        if k == "summary":      scache["summary"]    = do_gen_summary(selected_policy, combined)
                        elif k == "info":       scache["info"]       = do_gen_info(selected_policy, combined)
                        elif k == "mindmap":    scache["mindmap"]    = do_gen_mm(selected_policy, combined)
                        elif k == "flashcards": scache["flashcards"] = do_gen_fc(selected_policy, combined)
                        elif k == "quiz":       scache["quiz"]       = do_gen_qz(selected_policy, combined)
                        elif k == "report":     scache["report"]     = do_gen_report(selected_policy, all_src_text)
                        save_studio_cache(selected_policy, scache)
                    st.success(f"✅ {lbl} 저장 완료!"); st.rerun()

    # ── 콘텐츠 서브탭 (마인드맵 제외 — 관리자는 생성 패널에서 별도 관리) ──
    s1, s2, s3, s4 = st.tabs(["📋 요약 카드", "📊 인포그래픽", "🃏 플래시카드", "🧠 퀴즈 & 보고서"])

    with s1:
        cards = scache.get("summary")
        if cards: show_cards(cards)
        else: not_ready_msg()

    with s2:
        info = scache.get("info")
        if info: show_info(info)
        else: not_ready_msg()

    with s3:
        fc = scache.get("flashcards")
        if fc: show_fc(fc)
        else: not_ready_msg()

    with s4:
        ql, qr = st.columns(2)
        with ql:
            st.markdown("#### 🧠 퀴즈")
            qz = scache.get("quiz")
            if qz: show_qz(qz)
            else: not_ready_msg()
        with qr:
            st.markdown("#### 📄 보고서")
            report = scache.get("report")
            if report:
                st.markdown(report[:2000] + ("…" if len(report) > 2000 else ""))
                st.download_button("📥 보고서 (.md)", data=report,
                    file_name=f"{selected_policy}_보고서.md", mime="text/markdown", use_container_width=True)
            else: not_ready_msg()


# ════════════════════════════════════════════════════════════════
# 탭 3 ─ 정책 질의응답 챗봇
# ════════════════════════════════════════════════════════════════
with tab_chat:
    chat_col, src_col = st.columns([2.2, 1], gap="large")

    # ── 소스 패널 (오른쪽) ──────────────────────────────────
    with src_col:
        upload_srcs = [s for s in st.session_state.web_sources if s.get("type") == "pdf_upload"]
        web_srcs    = [s for s in st.session_state.web_sources if s.get("type") != "pdf_upload"]
        active_pdfs  = [f for f in pdfs if st.session_state.get(f"ck_{f}", True)]
        active_up    = [s for s in upload_srcs if st.session_state.get(f"ck_{s['id']}", True)]
        active_web   = [s for s in web_srcs    if st.session_state.get(f"ck_{s['id']}", True)]
        total_active = len(active_pdfs) + len(active_up) + len(active_web)

        st.markdown(f"""
<div style="background:#FAFAFA;border:1px solid #EEEEEE;border-radius:10px;
            padding:12px 14px;margin-bottom:8px;">
  <div style="font-weight:700;font-size:.9rem;color:#1A1A1A;margin-bottom:2px">
    📎 참고 소스
  </div>
  <div style="font-size:.78rem;color:#AAA">{total_active}개 선택됨</div>
</div>
""", unsafe_allow_html=True)

        ca, cb = st.columns(2)
        if ca.button("전체 선택", use_container_width=True, key="sel"):
            for f in pdfs: st.session_state[f"ck_{f}"] = True
            for s in st.session_state.web_sources: st.session_state[f"ck_{s['id']}"] = True
            st.rerun()
        if cb.button("전체 해제", use_container_width=True, key="desel"):
            for f in pdfs: st.session_state[f"ck_{f}"] = False
            for s in st.session_state.web_sources: st.session_state[f"ck_{s['id']}"] = False
            st.rerun()

        if pdfs:
            st.markdown('<div class="src-header">📄 PDF 문서</div>', unsafe_allow_html=True)
            for fname in pdfs:
                st.checkbox(fname[:26] + ("…" if len(fname) > 26 else ""),
                            key=f"ck_{fname}", value=st.session_state.get(f"ck_{fname}", True))
        if upload_srcs:
            st.markdown('<div class="src-header">📤 업로드 PDF</div>', unsafe_allow_html=True)
            for src in upload_srcs:
                st.checkbox(src["title"][:24] + ("…" if len(src["title"]) > 24 else ""),
                            key=f"ck_{src['id']}", value=st.session_state.get(f"ck_{src['id']}", True))
        if web_srcs:
            st.markdown('<div class="src-header">🌐 웹 & 유튜브</div>', unsafe_allow_html=True)
            for src in web_srcs:
                icon = "▶️" if src.get("type") == "youtube" else "📰"
                st.checkbox(f"{icon} {src['title'][:22]}{'…' if len(src['title'])>22 else ''}",
                            key=f"ck_{src['id']}", value=st.session_state.get(f"ck_{src['id']}", True))
        if not pdfs and not st.session_state.web_sources:
            st.info("소스가 없습니다.\n관리자가 추가합니다.")

    # ── 채팅 영역 (왼쪽) ────────────────────────────────────
    with chat_col:
        # 카카오톡 헤더
        today = datetime.datetime.now().strftime("%Y년 %m월 %d일")
        src_count = len(pdfs) + len(st.session_state.web_sources)
        st.markdown(f"""
<div class="kakao-header">
  <div class="k-avatar">🤖</div>
  <div>
    <div class="k-name">자치 매니저 톡톡이</div>
    <div class="k-sub">📂 {policy_display(selected_policy)} &nbsp;·&nbsp; 소스 {src_count}개</div>
  </div>
</div>
""", unsafe_allow_html=True)

        # 채팅 버블
        bubbles_html = '<div class="chat-wrap">'
        bubbles_html += f'<div class="chat-date-divider"><span>{today}</span></div>'
        if not st.session_state.messages:
            bubbles_html += '<div style="text-align:center;margin:auto;color:rgba(255,255,255,.7);font-size:.88rem;padding:30px 0">💬 아래에서 질문을 선택하거나<br>직접 입력해보세요</div>'
        for msg in st.session_state.messages:
            t = msg.get("time", "")
            if msg["role"] == "user":
                bubbles_html += f'<div class="bubble-user"><div class="time-left">{t}</div><div class="bubble">{msg["content"].replace(chr(10),"<br>")}</div></div>'
            else:
                content = msg["content"].replace(chr(10), "<br>")
                bubbles_html += f'<div class="bubble-ai"><div class="avatar">🤖</div><div class="bubble-body"><div class="sender">톡톡이</div><div class="bubble">{content}</div><div class="time-right">{t}</div></div></div>'
        bubbles_html += '</div>'
        # 입력창 영역 하단 라운딩용
        bubbles_html += '<div class="chat-input-area"></div>'
        st.markdown(bubbles_html, unsafe_allow_html=True)

        if total_active == 0:
            st.warning("⚠️ 오른쪽에서 참고 소스를 하나 이상 선택해주세요.")

        # 빠른 질문
        if not st.session_state.messages:
            st.markdown("**빠른 질문:**")
            examples = [
                "마을활동가 인정 방식을 지역별로 비교해주세요",
                "활동가 역량 기준은 어떻게 정의되나요?",
                "기회소득과 인정체계 연계 방안은?",
                "우수 사례와 정책 제언을 알려주세요",
            ]
            c1, c2 = st.columns(2)
            for i, q in enumerate(examples):
                if (c1 if i % 2 == 0 else c2).button(q, key=f"ex{i}", use_container_width=True):
                    if total_active == 0:
                        st.warning("소스를 먼저 선택해주세요.")
                    else:
                        st.session_state.messages.append({"role": "user", "content": q, "time": now_str()})
                        with st.spinner("답변 생성 중..."):
                            all_docs = {f: pdfs[f] for f in active_pdfs}
                            for s in active_up + active_web: all_docs[s["title"]] = s["text"]
                            ctx = get_chunks(q, all_docs)
                            ans = ask(q, ctx, st.session_state.messages)
                        st.session_state.messages.append({"role": "assistant", "content": ans, "time": now_str()})
                        st.rerun()

        # 입력창
        if prompt := st.chat_input("메시지를 입력하세요...", key="chat_input"):
            if total_active == 0:
                st.warning("소스를 먼저 선택해주세요.")
            else:
                st.session_state.messages.append({"role": "user", "content": prompt, "time": now_str()})
                with st.spinner("답변 생성 중..."):
                    all_docs = {f: pdfs[f] for f in active_pdfs}
                    for s in active_up + active_web: all_docs[s["title"]] = s["text"]
                    ctx = get_chunks(prompt, all_docs)
                    ans = ask(prompt, ctx, st.session_state.messages)
                st.session_state.messages.append({"role": "assistant", "content": ans, "time": now_str()})
                st.rerun()

        # 하단 버튼
        if st.session_state.messages:
            b1, b2 = st.columns(2)
            if b1.button("🗑️ 대화 초기화", key="clr", use_container_width=True):
                st.session_state.messages = []
                st.session_state.policy_note = ""
                st.rerun()
            if b2.button("📋 정책 노트 작성", key="make_note", use_container_width=True, type="primary"):
                if len(st.session_state.messages) < 2:
                    st.warning("대화를 조금 더 나눈 뒤 작성해주세요.")
                else:
                    hist_text = "\n".join(
                        f"{'[질문]' if m['role']=='user' else '[답변]'} {m['content']}"
                        for m in st.session_state.messages)
                    with st.spinner("정책 노트 작성 중..."):
                        note = call_gemini(f"""아래는 정책 공론장에서 나눈 대화입니다.
이 대화를 바탕으로 주민 누구나 이해할 수 있는 쉬운 용어로 정책 제안 노트를 작성해주세요.
반드시 아래 항목을 모두 포함하고 마크다운 형식으로 작성하세요:

## 주제
## 목적 (제안 취지)
## 관련 제도
## 정책 목표
## 실행 과제
### 문제점
### 해결 방안
## 소요 예산 (추정)
## 기대 효과

--- 대화 내용 ---
{hist_text[:20000]}
""")
                    st.session_state.policy_note = note
                    st.rerun()

        # 정책 노트 출력
        if st.session_state.get("policy_note"):
            st.markdown('<div class="note-area">', unsafe_allow_html=True)
            st.markdown("### 📋 정책 노트")
            st.markdown(st.session_state.policy_note)
            dl1, dl2 = st.columns(2)
            dl1.download_button("📥 노트 저장 (.md)", data=st.session_state.policy_note,
                file_name=f"{selected_policy}_정책노트.md", mime="text/markdown", use_container_width=True)
            if dl2.button("✕ 닫기", key="close_note", use_container_width=True):
                st.session_state.policy_note = ""; st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)


# ════════════════════════════════════════════════════════════════
# 소스 관리 (관리자 전용)
# ════════════════════════════════════════════════════════════════
if st.session_state.is_admin:
    st.divider()
    with st.expander("🗂️ 소스 관리 (관리자 전용)", expanded=False):
        st.markdown("### 🗂️ 소스 관리")
        st.caption("이곳에서 추가한 소스는 즉시 챗봇에 반영됩니다.")

        add1, add2, add3, add4 = st.tabs(["📤 PDF 업로드", "🔗 웹 / 뉴스 링크", "🎥 유튜브", "📊 엑셀 일괄 등록"])

        with add1:
            st.markdown("##### PDF 파일을 드래그하거나 선택하세요")
            uploaded_files = st.file_uploader("PDF 선택", type=["pdf"], accept_multiple_files=True,
                key="pdf_uploader", label_visibility="collapsed")
            if uploaded_files:
                existing_ids = {s["id"] for s in st.session_state.web_sources}
                new_count = 0
                for uf in uploaded_files:
                    fid = f"up_{abs(hash(uf.name + str(uf.size)))}"[:12]
                    if fid in existing_ids: continue
                    with st.spinner(f"'{uf.name}' 텍스트 추출 중..."):
                        try:
                            text = extract_pdf_bytes(uf.read())
                            if not text.strip():
                                st.warning(f"'{uf.name}' — 텍스트를 추출할 수 없습니다."); continue
                            ns = {"id": fid, "type": "pdf_upload", "title": uf.name, "url": "", "text": text[:30000]}
                            st.session_state.web_sources.append(ns)
                            st.session_state[f"ck_{fid}"] = True
                            existing_ids.add(fid); new_count += 1
                        except Exception as e: st.error(f"'{uf.name}' 오류: {e}")
                if new_count > 0:
                    save_sources(selected_policy, st.session_state.web_sources)
                    st.success(f"✅ {new_count}개 PDF 추가됨"); st.rerun()

        with add2:
            st.markdown("##### 웹사이트 또는 뉴스 기사 URL")
            u1, u2 = st.columns([4, 1])
            url_in = u1.text_input("URL", placeholder="https://...", key="url_in", label_visibility="collapsed")
            if u2.button("추가", key="btn_url", use_container_width=True):
                if url_in.strip().startswith("http"):
                    with st.spinner("읽는 중..."):
                        try:
                            title, text = fetch_article(url_in.strip())
                            fid = str(abs(hash(url_in)))[:10]
                            ns = {"id": fid, "type": "article", "title": title, "url": url_in.strip(), "text": text[:20000]}
                            st.session_state.web_sources.append(ns)
                            st.session_state[f"ck_{fid}"] = True
                            save_sources(selected_policy, st.session_state.web_sources)
                            st.success(f"✅ {title[:40]}"); st.rerun()
                        except Exception as e: st.error(f"실패: {e}")
                else: st.warning("https://... 형식으로 입력하세요.")
            st.divider()
            st.markdown("##### 키워드 검색")
            k1, k2 = st.columns([4, 1])
            kw_in = k1.text_input("검색어", placeholder="예: 마을활동가 인정체계", key="kw_in", label_visibility="collapsed")
            if k2.button("검색", key="btn_kw", use_container_width=True):
                with st.spinner("검색 중..."): st.session_state.search_res = search_web(kw_in)
            for i, r in enumerate(st.session_state.get("search_res", [])):
                with st.container(border=True):
                    rc1, rc2 = st.columns([5, 1])
                    with rc1:
                        st.markdown(f"**{r['title'][:50]}**"); st.caption(f"{r['snippet'][:100]}…"); st.caption(f"🔗 {r['url'][:60]}")
                    with rc2:
                        if st.button("추가", key=f"add_{i}", use_container_width=True):
                            with st.spinner("읽는 중..."):
                                try: title, text = fetch_article(r["url"])
                                except: title, text = r["title"], r["snippet"]
                                fid = str(abs(hash(r["url"])))[:10]
                                ns = {"id": fid, "type": "article", "title": title, "url": r["url"], "text": text[:20000]}
                                st.session_state.web_sources.append(ns)
                                st.session_state[f"ck_{fid}"] = True
                                save_sources(selected_policy, st.session_state.web_sources); st.rerun()

        with add3:
            st.markdown("##### 유튜브 자막 추출")
            y1, y2 = st.columns([4, 1])
            yt_in = y1.text_input("YouTube URL", placeholder="https://www.youtube.com/watch?v=...", key="yt_in", label_visibility="collapsed")
            if y2.button("추가", key="btn_yt", use_container_width=True):
                if "youtube.com" in yt_in or "youtu.be" in yt_in:
                    with st.spinner("자막 추출 중..."):
                        try:
                            title, text = fetch_youtube(yt_in.strip())
                            fid = str(abs(hash(yt_in)))[:10]
                            ns = {"id": fid, "type": "youtube", "title": title, "url": yt_in.strip(), "text": text[:20000]}
                            st.session_state.web_sources.append(ns)
                            st.session_state[f"ck_{fid}"] = True
                            save_sources(selected_policy, st.session_state.web_sources)
                            st.success(f"✅ {title[:40]}"); st.rerun()
                        except Exception as e: st.error(f"실패: {e}")
                else: st.warning("유튜브 URL을 입력하세요.")

        with add4:
            st.markdown("##### 엑셀 일괄 등록")
            if EXCEL_OK:
                import openpyxl
                wb_s = openpyxl.Workbook(); ws_s = wb_s.active; ws_s.title = "URL 목록"
                ws_s.append(["URL","제목(선택)","유형(선택)"])
                ws_s.append(["https://example.com","기사 제목","article"])
                ws_s.append(["https://www.youtube.com/watch?v=xxxxx","유튜브","youtube"])
                ws_s.column_dimensions['A'].width = 50; ws_s.column_dimensions['B'].width = 30
                buf2 = io.BytesIO(); wb_s.save(buf2)
                st.download_button("📥 엑셀 양식 다운로드", data=buf2.getvalue(),
                    file_name="소스_URL_양식.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            st.markdown("---")
            excel_file = st.file_uploader("엑셀 파일 선택", type=["xlsx"], key="excel_uploader", label_visibility="collapsed")
            if excel_file and EXCEL_OK:
                try:
                    wb2 = openpyxl.load_workbook(io.BytesIO(excel_file.read()))
                    ws2 = wb2.active; rows = list(ws2.iter_rows(values_only=True))
                    if len(rows) < 2: st.warning("2행부터 URL을 입력해주세요.")
                    else:
                        header = [str(c).strip().lower() if c else "" for c in rows[0]]
                        uc = next((i for i, h in enumerate(header) if "url" in h), 0)
                        tc = next((i for i, h in enumerate(header) if "제목" in h or "title" in h), None)
                        valid = [(r[uc], r[tc] if tc is not None else None)
                                 for r in rows[1:] if r[uc] and str(r[uc]).startswith("http")]
                        st.info(f"📋 {len(valid)}개 URL 감지")
                        for i, (url, title) in enumerate(valid[:20]):
                            icon = "▶️" if "youtube" in str(url) else "🔗"
                            st.markdown(f"`{i+1}` {icon} {str(url)[:70]}" + (f" — *{title}*" if title else ""))
                        if len(valid) > 20: st.caption(f"… 외 {len(valid)-20}개")
                        if st.button("✅ 전체 소스 추가 시작", type="primary", use_container_width=True, key="excel_add"):
                            existing_ids = {s["id"] for s in st.session_state.web_sources}
                            added, skipped, failed = 0, 0, 0
                            prog = st.progress(0, text="추가 중...")
                            for i, (url, hint_title) in enumerate(valid):
                                url = str(url).strip(); fid = str(abs(hash(url)))[:10]
                                if fid in existing_ids: skipped += 1
                                else:
                                    try:
                                        is_yt = "youtube.com" in url or "youtu.be" in url
                                        if is_yt: title, text = fetch_youtube(url); stype = "youtube"
                                        else: title, text = fetch_article(url); stype = "article"
                                        if hint_title and str(hint_title).strip(): title = str(hint_title).strip()
                                        ns = {"id": fid, "type": stype, "title": title, "url": url, "text": text[:20000]}
                                        st.session_state.web_sources.append(ns)
                                        st.session_state[f"ck_{fid}"] = True
                                        existing_ids.add(fid); added += 1
                                    except: failed += 1
                                prog.progress((i+1)/len(valid), text=f"{i+1}/{len(valid)} 처리 중")
                            save_sources(selected_policy, st.session_state.web_sources)
                            prog.empty(); st.success(f"✅ 추가 {added} · 스킵 {skipped} · 실패 {failed}"); st.rerun()
                except Exception as e: st.error(f"파일 읽기 오류: {e}")

        st.divider()
        st.markdown(f"### 📚 현재 소스 목록 — {policy_display(selected_policy)}")
        if pdfs:
            st.markdown(f"**📄 폴더 PDF** ({len(pdfs)}개)")
            for fname in pdfs:
                st.markdown(f'<div class="src-card"><div class="src-card-icon">📄</div><div class="src-card-body"><div class="src-card-title">{fname}</div><div class="src-card-meta">폴더 파일 · {len(pdfs[fname])//1000}K자</div><span class="src-card-badge badge-pdf">PDF</span></div></div>', unsafe_allow_html=True)
        if st.session_state.web_sources:
            st.markdown(f"**🌐 추가된 소스** ({len(st.session_state.web_sources)}개)")
            for src in list(st.session_state.web_sources):
                stype = src.get("type","article")
                url_disp = f'<a href="{src["url"]}" target="_blank">{src["url"][:50]}</a>' if src.get("url") else "업로드 파일"
                chars = len(src.get("text","")) // 1000
                cm, cd = st.columns([10, 1])
                with cm:
                    st.markdown(f'<div class="src-card"><div class="src-card-icon">{src_icon(stype)}</div><div class="src-card-body"><div class="src-card-title">{src["title"]}</div><div class="src-card-meta">{url_disp} · {chars}K자</div>{src_badge(stype)}</div></div>', unsafe_allow_html=True)
                with cd:
                    st.markdown("<div style='margin-top:10px'></div>", unsafe_allow_html=True)
                    if st.button("🗑️", key=f"del_{src['id']}", help="소스 삭제"):
                        st.session_state.web_sources = [s for s in st.session_state.web_sources if s["id"] != src["id"]]
                        save_sources(selected_policy, st.session_state.web_sources); st.rerun()
        else:
            st.info("추가된 소스가 없습니다.")


# ════════════════════════════════════════════════════════════════
# 푸터
# ════════════════════════════════════════════════════════════════
st.markdown("""
<div style="
  margin-top:44px;padding:16px 24px;
  border-top:1px solid #EEEEEE;
  background:#FAFAFA;border-radius:0 0 12px 12px;
  display:flex;justify-content:space-between;align-items:center;
  flex-wrap:wrap;gap:8px;font-size:.79rem;color:#AAA;line-height:1.7;
">
  <div>
    © 2026 위서영 &nbsp;/&nbsp;
    <a href="https://creativecommons.org/licenses/by/4.0/deed.ko"
       target="_blank" style="color:#888;text-decoration:none;">CC BY 4.0 (저작권표시)</a>
    &nbsp;·&nbsp; 위서영(위트콘텐츠연구소) &nbsp;·&nbsp;
    010-9887-9366 &nbsp;/&nbsp;
    <a href="mailto:witconlab@naver.com" style="color:#888;text-decoration:none;">witconlab@naver.com</a>
  </div>
  <div>💾 데이터는 메모리에만 유지 (새로고침 시 초기화)</div>
</div>
""", unsafe_allow_html=True)
